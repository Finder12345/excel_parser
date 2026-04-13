# -*- coding: utf-8 -*-
"""
读取 trace 指向的原始单元格及邻居值。

纯函数，零LLM。
"""
import json
from typing import Any, Dict, Optional, Union

import openpyxl
from langchain.tools import tool

from .column_utils import col_index_to_letter_core, col_letter_to_index_core


def _to_col(col: Union[int, str]) -> int:
    if isinstance(col, int):
        return col
    return col_letter_to_index_core(str(col))


def read_source_cell_core(file_path: str, sheet_name: str, row: int, col: Union[int, str]) -> Dict[str, Any]:
    c = _to_col(col)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]

    cell = ws.cell(row=row, column=c)
    raw = "" if cell.value is None else cell.value

    merged_from: Optional[str] = None
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= c <= mr.max_col:
            if not (mr.min_row == row and mr.min_col == c):
                merged_from = f"{col_index_to_letter_core(mr.min_col)}{mr.min_row}"
            break

    neighbors = {
        "above": "" if row <= 1 else ("" if ws.cell(row=row - 1, column=c).value is None else ws.cell(row=row - 1, column=c).value),
        "below": "" if ws.cell(row=row + 1, column=c).value is None else ws.cell(row=row + 1, column=c).value,
        "left": "" if c <= 1 else ("" if ws.cell(row=row, column=c - 1).value is None else ws.cell(row=row, column=c - 1).value),
        "right": "" if ws.cell(row=row, column=c + 1).value is None else ws.cell(row=row, column=c + 1).value,
    }

    wb.close()
    return {
        "raw_value": raw,
        "merged_from": merged_from,
        "neighbors": neighbors,
    }


@tool
def read_source_cell(file_path: str, sheet_name: str, row: int, col: str) -> str:
    """Read a source cell and neighboring context as JSON."""
    result = read_source_cell_core(file_path, sheet_name, row, col)
    return json.dumps(result, ensure_ascii=False, default=str)
