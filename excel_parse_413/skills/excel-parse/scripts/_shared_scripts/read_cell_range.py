# -*- coding: utf-8 -*-
"""
读取 Excel 指定范围单元格。

纯函数，零LLM。
"""
import json
from typing import Any, List, Union

import openpyxl
from langchain.tools import tool

from .column_utils import col_letter_to_index_core


def _to_col_index(col: Union[int, str]) -> int:
    if isinstance(col, int):
        return col
    return col_letter_to_index_core(str(col))


def read_cell_range_core(
    file_path: str,
    sheet_name: str,
    start_row: int,
    start_col: Union[int, str],
    end_row: int,
    end_col: Union[int, str],
) -> List[List[Any]]:
    """读取指定范围并返回二维数组。行列均为 1-based。"""
    s_col = _to_col_index(start_col)
    e_col = _to_col_index(end_col)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    out: List[List[Any]] = []
    for r in range(start_row, end_row + 1):
        row_vals: List[Any] = []
        for c in range(s_col, e_col + 1):
            val = ws.cell(row=r, column=c).value
            row_vals.append("" if val is None else val)
        out.append(row_vals)
    wb.close()
    return out


@tool
def read_cell_range(
    file_path: str,
    sheet_name: str,
    start_row: int,
    start_col: str,
    end_row: int,
    end_col: str,
) -> str:
    """Read a rectangular cell range from a worksheet."""
    data = read_cell_range_core(file_path, sheet_name, start_row, start_col, end_row, end_col)
    return json.dumps(data, ensure_ascii=False, default=str)
