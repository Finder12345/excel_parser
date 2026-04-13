# -*- coding: utf-8 -*-
"""
读取并清洗 Excel sheet 数据。

处理合并单元格展开、空行空列清理，返回干净的二维数组。

纯函数，零LLM。
"""
import json
from typing import Any, List, Optional, Union

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from langchain.tools import tool


# ── 核心纯函数 ──

def _handle_merged_cells(sheet: Worksheet) -> List[List[Any]]:
    """展开合并单元格，将顶左角的值填充到所有被合并的单元格。

    Args:
        sheet: openpyxl Worksheet 对象

    Returns:
        二维列表 [max_row][max_col]，合并单元格已展开，None 替换为 ""
    """
    max_row = sheet.max_row
    max_col = sheet.max_column
    if not max_row or not max_col:
        return []

    data = [[None for _ in range(max_col)] for _ in range(max_row)]

    # 先处理合并区域
    merged_ranges = list(sheet.merged_cells.ranges) if sheet.merged_cells else []
    merged_values = {}
    for merged in merged_ranges:
        top_left = sheet.cell(row=merged.min_row, column=merged.min_col)
        merged_value = "" if top_left.value is None else top_left.value
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                merged_values[(r, c)] = merged_value

    # 填充数据
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            key = (row, col)
            if key in merged_values:
                data[row - 1][col - 1] = merged_values[key]
            else:
                cell = sheet.cell(row=row, column=col)
                data[row - 1][col - 1] = "" if cell.value is None else cell.value
    return data


def _clean_sheet_data(sheet_data: List[List[Any]]) -> List[List[Any]]:
    """清理空行和空列。

    Args:
        sheet_data: 二维列表

    Returns:
        去除全空行和全空列后的二维列表
    """
    if not sheet_data:
        return []
    # 对齐原始 excel_parse 逻辑：
    # handle_merged_cells 已将空单元格转为 ""，这里只按 NaN 清理，
    # 不把 "" 视为缺失值，否则会改变 trace 的行/列定位。
    df = pd.DataFrame(sheet_data)
    df = df.dropna(how="all")  # 去空行
    df = df.dropna(how="all", axis=1)  # 去空列
    return df.values.tolist()


def read_sheet_core(
    file_path: str,
    sheet_name: str,
    max_rows: Optional[int] = None,
    max_cols: Optional[int] = None,
) -> List[List[Any]]:
    """读取并清洗 Excel sheet 数据。

    处理流程：
    1. 用 openpyxl 打开指定 sheet
    2. 展开所有合并单元格（填充顶左角值）
    3. 去除全空行和全空列
    4. 可选截取前 N 行/列

    Args:
        file_path: Excel 文件路径
        sheet_name: sheet 名称
        max_rows: 最大读取行数（None=全部）
        max_cols: 最大读取列数（None=全部）

    Returns:
        二维列表，合并单元格已展开，空行列已清理
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    sheet = wb[sheet_name]
    data = _handle_merged_cells(sheet)
    wb.close()

    data = _clean_sheet_data(data)

    # 截取
    if max_rows and len(data) > max_rows:
        data = data[:max_rows]
    if max_cols:
        data = [row[:max_cols] if len(row) > max_cols else row for row in data]

    return data


# ── LangChain @tool 包装 ──

@tool
def read_sheet(
    file_path: str,
    sheet_name: str,
    max_rows: Optional[int] = None,
    max_cols: Optional[int] = None,
) -> str:
    """读取并清洗 Excel sheet 数据。合并单元格已展开，空行列已清理。返回 JSON 格式二维数组。

    Args:
        file_path: Excel 文件路径
        sheet_name: sheet 名称
        max_rows: 最大读取行数（不传=全部读取）
        max_cols: 最大读取列数（不传=全部读取）
    """
    data = read_sheet_core(file_path, sheet_name, max_rows, max_cols)
    return json.dumps(data, ensure_ascii=False, default=str)
