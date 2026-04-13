# -*- coding: utf-8 -*-
"""Excel helpers (standalone)."""

from __future__ import annotations

from typing import List, Union

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet


def clean_sheet_data(sheet_data):
    df = pd.DataFrame(sheet_data)
    df.dropna(how='all', inplace=True)
    df.dropna(how='all', axis=1, inplace=True)
    return df.values.tolist()


def handle_merged_cells(sheet: Worksheet) -> List[List[Union[str, int, float, bool, None]]]:
    max_row = sheet.max_row
    max_col = sheet.max_column
    data = [[None for _ in range(max_col)] for _ in range(max_row)]

    merged_ranges = list(sheet.merged_cells.ranges) if sheet.merged_cells else []
    merged_values = {}
    for merged in merged_ranges:
        top_left = sheet.cell(row=merged.min_row, column=merged.min_col)
        merged_value = "" if top_left.value is None else top_left.value
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                merged_values[(r, c)] = merged_value

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            key = (row, col)
            if key in merged_values:
                data[row - 1][col - 1] = merged_values[key]
            else:
                cell = sheet.cell(row=row, column=col)
                data[row - 1][col - 1] = "" if cell.value is None else cell.value

    return data
