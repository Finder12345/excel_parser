# -*- coding: utf-8 -*-
"""
@File    : ExcelToJson.py
@Date    : 2025--08-19 11:49
@Desc    : Description of the file
@Author  : lei
"""
import logging
import traceback

import pandas as pd
from .global_var import wbparser
import json
from openpyxl.utils import get_column_letter


def dataframe_to_excel_json(df):
    """
    将DataFrame转换为指定JSON格式，正确处理换行符

    参数:
        df: pandas DataFrame对象

    返回:
        JSON字符串，格式为 [{列字母: [{index: 行号, value: 值}, ...]}, ...]
    """
    try:
        result = []

        # 处理列名（转换为Excel列字母）
        for col_idx in range(df.shape[1]):
            col_list = []
            excel_col = get_column_letter(col_idx + 1)  # 转换为Excel列字母(A,B,C)

            # 处理每列数据
            for row_idx in range(df.shape[0]):
                value = df.iat[row_idx, col_idx]
                excel_row = row_idx + 1  # 转换为Excel行号(从1开始)

                # 处理空值
                if pd.isna(value):
                    value = None

                # 处理换行符：确保保留实际换行
                if isinstance(value, str):
                    # 替换Windows换行符为统一格式
                    value = value.replace('\r\n', '\n')

                # 构建行对象
                cell_data = {
                    "index": excel_row,
                    "value": value
                }
                col_list.append(cell_data)

            # 添加到结果
            result.append({excel_col: col_list})

        # 自定义JSON编码器处理换行符
        class CustomJSONEncoder(json.JSONEncoder):
            def __init__(self, *args, **kwargs):
                kwargs['ensure_ascii'] = False  # 确保非ASCII字符正确显示
                super().__init__(*args, **kwargs)

            def encode(self, obj):
                # 重写encode方法处理换行符
                json_str = super().encode(obj)
                # 将转义的换行符替换为实际换行
                return json_str.replace('\\n', '\n').replace('\\r', '')

        return json.dumps(result, indent=2, cls=CustomJSONEncoder)
    except Exception as e:
        logging.error(traceback.format_exc())




def remove_last_empty_columns_iterative(df):
    """迭代方式删除 DataFrame 末尾的“全空列”。

    判空策略：
        - 若行数 > 1：忽略第 1 行（通常是标题/说明），检查第 2..n 行是否全为 NaN；
        - 否则：检查整列是否全为 NaN。

    返回:
        删除末尾空列后的新 DataFrame（通过 df = df.iloc[:, :-1] 截断）。
    """
    while not df.empty:
        last_col_index = df.shape[1] - 1
        if df.shape[0] > 1:
            is_all_null = df.iloc[1:, last_col_index].isnull().all()
        else:
            is_all_null = df.iloc[:, last_col_index].isnull().all()

        if is_all_null:
            df = df.iloc[:, :-1]
        else:
            break
    return df


def remove_last_empty_rows_iterative(df):
    """迭代方式删除 DataFrame 末尾的“全空行”。

    判空策略：
        - 若列数 > 1：忽略第 1 列（通常是标题/索引列），检查第 2..m 列是否全为 NaN；
        - 否则：检查整行是否全为 NaN。

    返回:
        删除末尾空行后的新 DataFrame（通过 df = df.iloc[:-1, :] 截断）。
    """
    while not df.empty:
        last_row_index = df.shape[0] - 1
        if df.shape[1] > 1:
            is_all_null = df.iloc[last_row_index, 1:].isnull().all()
        else:
            is_all_null = df.iloc[last_row_index, :].isnull().all()

        if is_all_null:
            df = df.iloc[:-1, :]
        else:
            break
    return df


def excel_to_json(excel_path, sheet_name="GWRoutingChart"):
    """将 Excel 的指定 sheet 转为简化 JSON（用于 LLM 表头结构识别）。

    设计目的：
        仅抽取“前 10 行”的内容，降低 token 与模型负担，用于识别：
        - 表头行数/合并单元格结构
        - 关键字段列（Signal/PDU/矩阵区）的大致位置

    主要流程：
        1) 使用 wbparser 打开工作簿并选择 sheet（优先 sheet_name，否则 active）；
        2) 通过 sheet.values 构建 DataFrame；
        3) 仅处理前 10 行内的合并单元格：用左上角值填充合并区域；
        4) remove_last_empty_columns_iterative / remove_last_empty_rows_iterative 去掉末尾空列/空行；
        5) dataframe_to_excel_json(new_df1.head(10)) 输出 JSON 字符串。

    返回:
        (json_output, row_count, column_count)
        - json_output: 供 LLM 使用的 JSON 字符串
        - row_count/column_count: 处理后 DataFrame 的尺寸（用于后续遍历/坐标计算）
    """
    # 加载 Excel 文件
    workbook = wbparser.parse_and_get_single_excel(excel_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active  # 获取指定工作表或默认第一个工作表

    # 用 Pandas 读取 Excel 内容
    df = pd.DataFrame(sheet.values)

    # 获取合并单元格范围
    merged_ranges = sheet.merged_cells.ranges

    # 筛选出只在前10行的合并单元格范围
    merged_in_first_10_rows = [
        merged_range for merged_range in merged_ranges
        if merged_range.min_row <= 10
    ]

    # 遍历合并单元格范围
    for merged_range in merged_in_first_10_rows:
        # 获取合并单元格的起始值
        merged_value = sheet[merged_range.start_cell.coordinate].value

        # 填充合并单元格范围内的所有单元格
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                df.iloc[row - 1, col - 1] = merged_value  # 注意 Pandas 的行列索引从 0 开始

    # 转换为目标JSON格式
    new_df = remove_last_empty_columns_iterative(df)
    new_df1 = remove_last_empty_rows_iterative(new_df)
    json_output = dataframe_to_excel_json(new_df1.head(10))
    return json_output, new_df1.shape[0], new_df1.shape[1]

if __name__ == '__main__':
    file_paths = [
        r"D:\MyCode\Autosar-Agent-Temp\data\A02-Y_纯电AB平台_CMX_CCU_V29.2(G3 G.0-1)-20251201_Result.xlsx"]

    json_output, row_count, column_count = excel_to_json(file_paths[0],"CZL_CANFD DebugMessage")
    print(json_output)
    print(f"行数: {row_count}")
    print(f"列数: {column_count}")
