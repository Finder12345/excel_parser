# -*- coding: utf-8 -*-
import json
import logging
import os
import time
import traceback
from datetime import datetime
from langchain.chat_models import init_chat_model
import openpyxl
from pathlib import Path

from .dep_llm import get_models as _dep_get_models
from .global_var import input_modules
from .prompt import get_prompts_dict
# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
from .dep_llm import call_model

class WorkbookParser:
    def __init__(self):
        self.workbook_data = {}

    def parse_excel(self, file_paths):
        if not isinstance(file_paths, list):
            file_paths = [file_paths]

        for file_path in file_paths:
            if not os.path.exists(file_path):
                return
            if not file_path.endswith(('.xlsx', '.xls', '.xlsm')):
                continue
            if file_path in self.workbook_data:
                continue
            workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
            self.workbook_data[file_path] = workbook

    def parse_and_get_single_excel(self, file_path):
        if not os.path.exists(file_path):
            return
        if not file_path.endswith(('.xlsx', '.xls', '.xlsm')):
            return
        if file_path in self.workbook_data:
            return self.workbook_data[file_path]

        workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        self.workbook_data[file_path] = workbook
        return workbook


wbparser = WorkbookParser()

# 全局变量
CAN_MODULE = "CAN"
LIN_MODULE = "LIN"
ETH_MODULE = "ETH"
NVM_MODULE = "NVM"
DIAG_MODULE = "DIAG"
# input_modules = [CAN_MODULE, LIN_MODULE, NVM_MODULE]  # 默认包含这些模块
input_modules = []


# 知识库
class KnowledgeBase:
    def __init__(self):
        self.kb = self.get_knowledge()

    def get_knowledge(self):
        try:
            # 构建knowledge.json文件的路径
            knowledge_file_path = Path(__file__).resolve().parent/"knowledge.json"
            # knowledge_file_path = Path(r"D:\KoTEI_CODE\Agent_Autosar_Backend\AutoSAR_Agent\tools\OMV_tools\config\knowledge.json")
            # 如果文件不存在，返回默认值
            if not knowledge_file_path.exists():
                return self.get_default_knowledge()

            # 读取知识库文件
            with open(knowledge_file_path, 'r', encoding='utf-8') as f:
                knowledge_data = json.load(f)

            return knowledge_data
        except Exception:
            # 离线/本地运行允许缺失 knowledge.json；返回默认知识库
            return self.get_default_knowledge()

    def get_default_knowledge(self):
        # 返回默认知识库
        return {
            "sheet_info": {
                "can_routing": {
                    "sheet_name_key_words": ["路由", "routing", "gateway"],
                    "sheet_description": ["包含CAN消息路由信息的表格", "Gateway routing table"]
                },
                "can_signal": {
                    "sheet_name_key_words": ["信号", "signal", "CAN"],
                    "sheet_description": ["包含CAN信号定义的表格", "CAN signal definition table"]
                },
                "lin_signal": {
                    "sheet_name_key_words": ["LIN", "lin"],
                    "sheet_description": ["包含LIN信号定义的表格", "LIN signal definition table"]
                },
                "lin_schedule": {
                    "sheet_name_key_words": ["调度", "schedule", "LIN"],
                    "sheet_description": ["包含LIN调度信息的表格", "LIN schedule table"]
                },
                "nvm_block": {
                    "sheet_name_key_words": ["NVM", "存储", "memory"],
                    "sheet_description": ["包含存储配置信息的表格", "NVM configuration table"]
                }
            }
        }


kb = KnowledgeBase().kb

# 提示词数据
# prompts_data = {
#     "QueryGwSheetNamePrompt": "请从以下Excel工作表名称中，选择最可能包含{module_name}信息的工作表名称。\n"
#                               "关键词提示: {sheet_name_key_words}\n"
#                               "表格描述: {sheet_description}\n"
#                               "工作表名称列表: {{sheet_names}}\n"
#                               "请只返回工作表名称，不要返回其他内容。",
#     "QueryNvmSheetNamePrompt": "请从以下Excel工作表名称中，选择最可能包含{module_name}信息的工作表名称。\n"
#                                "关键词提示: {sheet_name_key_words}\n"
#                                "表格描述: {sheet_description}\n"
#                                "工作表名称列表: {{sheet_names}}\n"
#                                "请只返回工作表名称，不要返回其他内容。",
#     "LinSignalSheetNamePrompt": "请从以下Excel工作表名称中，选择最可能包含{module_name}信息的工作表名称。\n"
#                                 "关键词提示: {sheet_name_key_words}\n"
#                                 "表格描述: {sheet_description}\n"
#                                 "工作表名称列表: {{sheet_names}}\n"
#                                 "请返回所有可能的工作表名称，用逗号分隔。",
#     "LinScheduleSheetNamePrompt": "请从以下Excel工作表名称中，选择最可能包含{module_name}信息的工作表名称。\n"
#                                   "关键词提示: {sheet_name_key_words}\n"
#                                   "表格描述: {sheet_description}\n"
#                                   "工作表名称列表: {{sheet_names}}\n"
#                                   "请返回所有可能的工作表名称，用逗号分隔。",
#     "CanSignalSheetNamePrompt": "请从以下Excel工作表名称中，选择最可能包含{module_name}信息的工作表名称。\n"
#                                 "关键词提示: {sheet_name_key_words}\n"
#                                 "表格描述: {sheet_description}\n"
#                                 "工作表名称列表: {{sheet_names}}\n"
#                                 "请返回所有可能的工作表名称，用逗号分隔。"
# }
prompts_data = get_prompts_dict()


# 全局配置
class GlobalConfig:
    def __init__(self):
        self.file_type_info = {}
        self.current_work_space = {
            'project_directory': os.path.join(os.path.dirname(__file__), "output"),
            "project_name": "default"
        }
        self.parse_start_time = 0
        self.llm_spend_token = {'excel_parse': 0}
        self.task_dict = {}
        self.previous_data = {}
        self.excel_exception_list = []
        self.channel_trace_info = {}


# 初始化全局配置
global_config = GlobalConfig()


# LLM 模型 - 使用 dep_llm 中已有的模型，不再依赖外部 init_llm
# 保留 get_models 函数签名兼容性
def get_models():
    """获取LLM模型 - 委托给 dep_llm"""
    return _dep_get_models()


# 获取sheet名称
# def get_sheet_name(sheet_names, prompt, module_name):
#     """使用LLM获取sheet名称"""
#     try:
#         # 构建提示词
#         prompt = prompt.format(sheet_names=", ".join(sheet_names))
#
#         # 获取模型
#         model = get_models().get("cur_chat")
#         if not model or not model.model:
#             # 如果模型未初始化，使用默认逻辑
#             logging.warning("LLM模型未初始化，使用默认逻辑")
#             # 默认逻辑：基于关键词匹配
#             if module_name == "Lin信号表":
#                 # 查找包含LIN的工作表
#                 lin_sheets = [name for name in sheet_names if "LIN" in name.upper()]
#                 if lin_sheets:
#                     return ", ".join(lin_sheets)
#             return ""
#
#         # 构建消息
#         messages = [
#             {"role": "system", "content": "你是一个Excel文件分析助手，擅长识别工作表的功能。"},
#             {"role": "user", "content": prompt}
#         ]
#
#         # 调用模型
#         response = model.invoke(messages)
#
#         # 提取内容
#         if hasattr(response, "content"):
#             sheet_name = response.content.strip()
#         else:
#             sheet_name = str(response).strip()
#
#         # 处理多个sheet名称的情况
#         sheet_names_list = [name.strip() for name in sheet_name.split(",")]
#         valid_sheet_names = []
#         for name in sheet_names_list:
#             if name in sheet_names:
#                 valid_sheet_names.append(name)
#             else:
#                 # 尝试模糊匹配
#                 for sheet in sheet_names:
#                     if name.lower() in sheet.lower():
#                         valid_sheet_names.append(sheet)
#                         break
#
#         if valid_sheet_names:
#             return ", ".join(valid_sheet_names)
#         else:
#             return ""
#     except Exception as e:
#         logging.error(f"获取sheet名称失败: {str(e)}")
#         return ""

def get_sheet_name(sheet_names, get_sheet_name_prompt, table):
    import json
    import re
    from json_repair import loads, repair_json

    def extract_json(content):
        try:
            return json.loads(content)
        except Exception:
            pass

        pattern = r"```json(.*?)```"
        matches = re.findall(pattern, content, re.DOTALL)
        if matches:
            json_content = matches[0].strip()
            json_content = json_content.replace("None", "null").replace("True", "true").replace("False", "false")
            return json.loads(json_content)

        processed_content = content.replace("None", "null").replace("True", "true").replace("False", "false")
        repaired = repair_json(processed_content)
        return loads(repaired)

    query = f"""
          请从以下名称列表列表中，找到表示{table}的名称：
          {sheet_names}
          """

    res = call_model(
        query.replace("{", "{{").replace("}", "}}"),
        get_sheet_name_prompt.replace("{", "{{").replace("}", "}}"),
    )

    # dep_llm.call_model 已不返回 token；兼容这里的统计逻辑
    spend_token = 0

    if isinstance(res, tuple):
        res, spend_token = res

    # global_config.llm_spend_token["excel_parse"] += spend_token

    res_json = str(res).replace("{{", "{").replace("}}", "}")
    sheet_name = extract_json(res_json).get("name", "")
    return sheet_name



# 通过AI获取sheet名称
def get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names):
    sheet_name_key_words = kb.get("sheet_info", {}).get(module, {}).get("sheet_name_key_words", [])
    sheet_name_key_words_str = '、'.join(sheet_name_key_words)
    sheet_description = kb.get("sheet_info", {}).get(module, {}).get("sheet_description", [])
    sheet_description_str = '。'.join(sheet_description) + "。"
    generate_prompt = input_prompt.format(
        sheet_name_key_words=sheet_name_key_words_str,
        sheet_description=sheet_description_str,
        module_name=module_name)
    sheet_name = get_sheet_name(sheet_names, generate_prompt, module_name)
    return sheet_name


# 获取CAN路由表名称
def get_can_routing_name(file_paths):
    res = []
    for file_path in file_paths:
        res1 = {
            "routing_name": "",
            "file_path": ""
        }
        wb = wbparser.parse_and_get_single_excel(file_path)
        if not wb:
            logging.error(f"Failed to parse Excel file: {file_path}")
            res.append(res1)
            continue
        sheet_names = wb.sheetnames
        input_prompt = prompts_data.get("QueryGwSheetNamePrompt")
        module = "can_routing"
        module_name = "网关路由表"
        gw_routing_name = get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names)
        res1["routing_name"] = gw_routing_name
        res1["file_path"] = file_path
        res.append(res1)
    return res


# 获取CAN信号表名称
def get_can_signal_sheet_names(file_paths):
    res = []
    for file_path in file_paths:
        workbook = wbparser.parse_and_get_single_excel(file_path)
        if not workbook:
            logging.error(f"Failed to parse Excel file: {file_path}")
            continue
        sheet_names = workbook.sheetnames
        input_prompt = prompts_data.get("CanSignalSheetNamePrompt")
        module = "can_signal"
        module_name = "CAN信号表"
        check_sheet_names = get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names)
        if check_sheet_names:
            # 处理可能的多个sheet名称
            # sheet_names_list = [name.strip() for name in check_sheet_names.split(",")]
            res.append({"file_path": file_path, "sheet_names": check_sheet_names})
    return res


# 获取NVM表名称
def get_nvm_sheet_name(file_paths):
    res = []
    for file_path in file_paths:
        wb = wbparser.parse_and_get_single_excel(file_path)
        if not wb:
            logging.error(f"Failed to parse Excel file: {file_path}")
            res.append({"sheet_name": "", "file_path": file_path})
            continue
        sheet_names = wb.sheetnames
        input_prompt = prompts_data.get("QueryNvmSheetNamePrompt")
        module = "nvm_block"
        module_name = "存储配置表"
        sheet_name = get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names)
        res.append({"sheet_name": sheet_name, "file_path": file_path})
    return res


# 获取LIN信号和调度表数据
def get_lin_signal_schedule_data(file_paths):
    res = []
    for file_path in file_paths:
        workbook = wbparser.parse_and_get_single_excel(file_path)
        if not workbook:
            logging.error(f"Failed to parse Excel file: {file_path}")
            continue
        sheet_names = workbook.sheetnames
        input_prompt = prompts_data.get("LinSignalSheetNamePrompt")
        module = "lin_signal"
        module_name = "Lin信号表"
        lin_sheet_names = get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names)
        if lin_sheet_names:
        #     sheet_names_list = [name.strip() for name in lin_sheet_names.split(",")]
            res.append({"file_path": file_path, "sheet_names": lin_sheet_names, 'sheet_type': 'lin'})
        input_prompt = prompts_data.get("LinScheduleSheetNamePrompt")
        module = "lin_schedule"
        module_name = "Lin调度表"
        schedule_sheet_names = get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names)
        # if schedule_sheet_names:
            # sheet_names_list = [name.strip() for name in schedule_sheet_names.split(",")]
        res.append({"file_path": file_path, "sheet_names": schedule_sheet_names, 'sheet_type': 'schedule'})
    return res


# 验证CAN路由数据
def validate_can_routing_data(can_routing_data):
    if CAN_MODULE not in input_modules:
        logging.info(f"CAN module not selected, skipping CAN module data validation")
        return True
    for can_routing_data_item in can_routing_data:
        if can_routing_data_item.get('routing_name'):
            return True
    else:
        logging.info(f"CAN module data is empty")
        return False


# 验证NVM数据
def validate_nvm_data(nvm_data):
    if NVM_MODULE not in input_modules:
        logging.info(f"NVM module not selected, skipping NVM module data validation")
        return True
    for nvm_data_item in nvm_data:
        if nvm_data_item.get('sheet_name'):
            return True
    else:
        logging.info(f"NVM module data is empty")
        return False


# 验证LIN数据
def validate_lin_data(lin_signal_scheduler_data):
    if LIN_MODULE not in input_modules:
        logging.info(f"LIN module not selected, skipping LIN module data validation")
        return True
    for lin_signal_scheduler_data_item in lin_signal_scheduler_data:
        if lin_signal_scheduler_data_item.get('sheet_names') and lin_signal_scheduler_data_item.get(
                'sheet_type') == 'lin':
            return True
    else:
        logging.info(f"LIN module data is empty")
        return False


msg_template = '''{module}模块数据为空。1.请首先确保导入需求时输入文件与选择的模块匹配一致！如果不一致请重新导入需求
                2.如果您的输入文件与选择的模块是一致的，那请告诉我正确的"模块-sheet_name"映射信息，我将更新到知识库，更新成功后您可以重新执行工作流。示例输入:LIN模块对应sheet_name是LIN1和LIN2'''


def extract_sheet_range(global_config:GlobalConfig):
    logging.info(
        f"**Start intelligent parsing of requirement file, generate overall reading rules for requirement file**")
    logging.info(f"[performance test] excel parsing started")
    excel_paths = global_config.file_type_info.get('excel_path')
    if excel_paths:
        try:
            global_config.parse_start_time = time.time()
            parse_start_time = global_config.parse_start_time
            global_config.llm_spend_token['excel_parse'] = 0
            # 避免存储无法 JSON 序列化的对象，只存储模型名称
            cur_model = get_models().get("cur_chat")
            model_name = getattr(cur_model, "model_name", str(cur_model)) if cur_model else "unknown"
            global_config.task_dict[int(parse_start_time)] = {
                "file_path": global_config.file_type_info.get("excel_path"),
                "task_type": "parse_excel",
                "spend_token": global_config.llm_spend_token['excel_parse'],
                "model": model_name
            }
            # 生成can相关待确认信息
            can_routing_data = get_can_routing_name(excel_paths)
            # print(f"sa:{can_routing_data}")
            if not validate_can_routing_data(can_routing_data):
                logging.error("需求文件的整体读取规则生成失败，失败信息：CAN模块数据为空")
                return {
                    "code": 0,
                    "msg": msg_template.format(module='CAN'),
                }


            can_signal_sheet_data = get_can_signal_sheet_names(excel_paths)
            nvm_data = get_nvm_sheet_name(excel_paths)
            if not validate_nvm_data(nvm_data):
                logging.error("需求文件的整体读取规则生成失败，失败信息：NVM模块数据为空")
                return {
                    "code": 0,
                    "msg": msg_template.format(module='NVM'),
                }
            # 生成lin相关待确认信息
            lin_signal_scheduler_data = get_lin_signal_schedule_data(excel_paths)
            if not validate_lin_data(lin_signal_scheduler_data):
                logging.error("需求文件的整体读取规则生成失败，失败信息：LIN模块数据为空")
                return {
                    "code": 0,
                    "msg": msg_template.format(module='LIN'),
                }
            # 获取json输出的文件夹
            output_dir = global_config.current_work_space['project_directory']
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            # 获取json输出路径
            project_name = global_config.current_work_space.get("project_name", "default")
            project_dir = os.path.join(output_dir, project_name)
            if not os.path.exists(project_dir):
                os.makedirs(project_dir)

            extract_routing_data_and_signal_sheet_data_path = os.path.join(
                project_dir,
                "extract_routing_data_and_signal_sheet_data.json"
            )
            with open(extract_routing_data_and_signal_sheet_data_path, 'w', encoding='utf-8') as f:
                requirement_metadata = {
                    "can_routing_data": can_routing_data,
                    "can_signal_sheet_data": can_signal_sheet_data,
                    "lin_signal_scheduler_data": lin_signal_scheduler_data,
                    "nvm_data": nvm_data,
                }
                json.dump(requirement_metadata, f, ensure_ascii=False, indent=2)
            global_config.previous_data[
                "extract_routing_data_and_signal_sheet_data"] = extract_routing_data_and_signal_sheet_data_path
            logging.info(
                f"**Successfully extracted overall reading rules for requirement file, path: {extract_routing_data_and_signal_sheet_data_path}**"
            )
            return extract_routing_data_and_signal_sheet_data_path
            # return {
            #     "code": 1,
            #     "msg": '',
            # }
        except Exception as e:
            error_msg = f"""Failed to generate overall reading rules for requirement file, error information: {traceback.format_exc()}"""
            logging.info(error_msg)
            # return {
            #     "code": 0,
            #     "msg": error_msg,
            # }
    else:
        return {
            "code": 0,
            "msg": '无excel文件',
        }


if __name__ == "__main__":
    # 配置
    global_config.file_type_info['excel_path'] = [
        r"D:\KoTEI_CODE\Agent_Autosar_Backend\data\EEA3.0混动系列A（适用T09-N等）_CMX_CCU_LIN_MIX_V1.1_20250326（格式修正）.xlsx",
        # r"D:\KoTEI_CODE\Agent_Autosar_Backend\data\CCU_MID_NVM(┤µ┤ó╣▄└φ╓╨╝Σ╝■─ú┐Θ)_BlockList_A664WD.xlsx",
        r"D:\KoTEI_CODE\Agent_Autosar_Backend\data\A02-Y_纯电AB平台_CMX_CCU_V29.2(G3 G.0-1)-20251201_Result.xlsx"]
    # global_config.file_type_info['excel_path'] = [r"D:\KoTEI_CODE\Agent_Autosar_Backend\new_data\AY5-G Project_CMX_EEA3.0_CCU_LIN_MIX_V1.4 - 增加诊断(LIN诊断相关).xlsx"]
    # global_config.file_type_info['excel_path'] = [
    #     r"D:\code\Agent_Autosar_Backend\data\EEA3.0混动系列A（适用T09-N等）_CMX_CCU_LIN_MIX_V1.1_20250326（格式修正）.xlsx",
    #     r"D:\KoTEI_CODE\Agent_Autosar_Backend\data\CCU_MID_NVM(┤µ┤ó╣▄└φ╓╨╝Σ╝■─ú┐Θ)_BlockList_A664WD.xlsx",
    # ]
    global_config.current_work_space['project_directory'] = r"D:\KoTEI_CODE\Agent_Autosar_Backend\data\output"
    global_config.current_work_space["project_name"] = "测试"

    # 运行
    result = extract_sheet_range(None)
    print(result)