# -*- coding: utf-8 -*-
"""
@File    : excel_ai_parse.py
@Date    : 2025--07-29 20:43
@Desc    : Description of the file
@Author  : lei
"""
import json
import os



from .common import get_sheet_name
from .dbc_parser_tool import DbcParser
from .parse_signal import parse_all_signal, schedule_parse_function, get_header_input, \
    nvm_map_function, table_header_map_function, get_excel_column_mapping, map_function
from .logic_LLM import parse_can_excel
from .parse_excel_api import parse_routing_table_api
from .global_var import wbparser
from . import global_var
from .global_cfg import global_config
import logging
import pandas as pd
from .prompt import prompts_data
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Union

def handle_merged_cells(sheet: Worksheet) -> List[List[str]]:
    """
    处理合并单元格（修正版）
    改进点：
    1. 保留原始0/False等值（原逻辑会错误转成空字符串）
    2. 显式处理None值
    3. 优化合并区域判断性能
    """
    max_row = sheet.max_row
    max_col = sheet.max_column
    data: List[List[Union[str, int, float, bool, None]]] = [[None for _ in range(max_col)] for _ in range(max_row)]

    # 预计算合并区域信息
    merged_ranges = list(sheet.merged_cells.ranges) if sheet.merged_cells else []

    # 构建合并单元格映射，键为(行,列)，值为合并区域左上角的值
    merged_values = {}
    for merged in merged_ranges:
        top_left = sheet.cell(row=merged.min_row, column=merged.min_col)
        merged_value = "" if top_left.value is None else top_left.value
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                merged_values[(r, c)] = merged_value

    # 一次性处理所有单元格
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell_key = (row, col)
            if cell_key in merged_values:
                # 合并单元格，使用预计算的值
                data[row - 1][col - 1] = merged_values[cell_key]
            else:
                # 非合并单元格，直接获取值
                cell = sheet.cell(row=row, column=col)
                data[row - 1][col - 1] = "" if cell.value is None else cell.value

    return data


def clean_sheet_data(sheet_data):
    """
    清理Sheet数据，去除空行和空列。

    :param sheet_data: List[List]，表示Excel表格的二维数据
    :return: 清理后的数据
    """
    # 将数据转换为DataFrame（方便处理）
    df = pd.DataFrame(sheet_data)

    # 去除空行（所有列均为空的行）
    df.dropna(how='all', inplace=True)

    # 去除空列（所有行均为空的列）
    df.dropna(how='all', axis=1, inplace=True)

    # 将清理后的数据转换回List[List]格式
    cleaned_data = df.values.tolist()
    return cleaned_data

class AIParseExcel(object):

    def __init__(self, file_path):
        """Excel 解析编排器（路由表/信号/NVM/LIN 等能力的组合入口）。

        这个类主要承担"流程编排"角色：
        - 根据全局上传文件类型（excel/dbc）选择不同的解析路径；
        - 先用 LLM/规则确定关键 sheet（如网关路由表）；
        - 再调用下游解析器输出结构化中间结果（excel_requirement_metadata_file.json）。

        参数:
            file_path: Excel 文件绝对路径。
        """
        self.file_path = file_path

    def get_routing_msg(self):
        """解析并返回路由表的结构化"路由表头/关键列"信息。

        主要流程：
        - 打开工作簿，获取全部 sheet 名称；
        - 使用 prompts_data 中的 QueryGwSheetNamePrompt 让模型/规则选择"网关路由表"sheet；
        - 调用 routingtable.logic_LLM.parse_can_excel 对该 sheet 做结构识别，得到 routing_msg。

        返回:
            (routing_msg, gw_routing_name)
            - routing_msg: 路由表结构识别结果（表头行数、关键列坐标、矩阵区含义等）。
            - gw_routing_name: 被识别为网关路由表的 sheet 名称。

        说明:
            这里返回的是"结构元数据"，真正的逐行扫描与一对多归并在 parse_routing_table_api/logic_AutoPy 中完成。
        """
        routing_msg = {}
        gw_routing_name = ""
        # 如果用户没有上传了dbc ，则使用excel解析信号
        if global_config.file_type_info.get('excel_path'):
            # 解读路由表头
            workbook = wbparser.parse_and_get_single_excel(self.file_path)
            sheet_names = workbook.sheetnames
            excel_query_gw_sheet_prompt = prompts_data.get("QueryGwSheetNamePrompt")
            gw_routing_name = get_sheet_name(sheet_names, excel_query_gw_sheet_prompt, "网关路由表")
            routing_msg = parse_can_excel(self.file_path, gw_routing_name)
        return routing_msg, gw_routing_name

    def extract_requirement_metadata(self, routing_msg, gw_routing_name):
        """生成"Excel 需求元数据"JSON 文件（供后续生成 Structured Content 使用）。

        主要流程：
        1) 先清空本轮解析的全局状态：
           - global_config.excel_exception_list：记录识别失败/需要用户纠错的项；
           - global_config.channel_trace_info：记录通道名推断与 trace 信息。
        2) 信号/报文来源分流：
           - 如果上传了 DBC：通过 DbcParser 解析得到 msg_info（信号数据可能来自 DBC/或此处为空）；
           - 否则：走 Excel 信号解析 parse_all_signal（注意：当前版本可能未启用，返回空 dict）。
        3) 如果存在路由表 sheet：调用 parse_routing_table_api 扫描并转换路由信息：
           - gw_mapping：网关映射（通常是 source -> [targets...] 的一对多结构 + trace）；
           - route_info：逐行抽取到的 route 列表（包含 routingType 等）。
        4) 将 signal/msg/gw/route 四类信息写入 excel_requirement_metadata_file.json。

        参数:
            routing_msg: 来自 get_routing_msg 的路由表结构识别结果。
            gw_routing_name: 网关路由表 sheet 名称。

        返回:
            excel_requirement_metadata_file: 生成的 JSON 文件路径。
        """
        # 调用之前先制空
        global_config.excel_exception_list = []
        global_config.channel_trace_info = {}
        gw_mapping = {}
        route_info = []
        # 如果用户上传了dbc ，则使用dbc 解析信号
        if global_config.file_type_info.get('dbc_info'):
            dbc_obj = DbcParser()

            msg_info = dbc_obj.parse_dbcs(global_config.file_type_info['dbc_info'])
            signal_datas = {}

        else:
            signal_datas = parse_all_signal(self.file_path)
            msg_info = {}
        if self.file_path and gw_routing_name:
            gw_mapping, route_info = parse_routing_table_api(self.file_path, routing_msg, gw_routing_name)
        # 获取json输出的文件夹
        output_dir = global_config.current_work_space['project_directory']
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        # 获取json输出路径
        excel_requirement_metadata_file = os.path.join(
            output_dir,
            str(global_config.current_work_space['project_name']),
            "excel_requirement_metadata_file.json"
        )
        # 写入json文件
        with open(excel_requirement_metadata_file, 'w', encoding='utf-8') as f:
            requirement_metadata = {"signal_datas": signal_datas, "msg_info": msg_info, "gw_mapping": gw_mapping,
                                    "route_info": route_info}
            json.dump(requirement_metadata, f, ensure_ascii=False, indent=2)
        return excel_requirement_metadata_file

    def update_srs_json(self, excel_requirement_metadata_file):
        """将 excel_requirement_metadata_file 的内容进一步更新到 Structured Content（占位）。

        说明:
            该方法目前为占位（pass）。从注释看，历史上计划：
            - 读取 excel_requirement_metadata_file.json；
            - 组装为 arxml_required_data/中间结构；
            - 调用 DataAssemblyForExcel 等组件生成 Structured Content。

        参数:
            excel_requirement_metadata_file: extract_requirement_metadata 生成的元数据 JSON 文件路径。
        """
        # 获取arxml_required_data数据
        # with open(excel_requirement_metadata_file, 'r', encoding='utf-8') as f:
        #     excel_requirement_metadata = json.load(f)
        # data = assemble_all(excel_requirement_metadata)
        # from app.eb.parse_req.parse_excel.data_assembly_for_excel import DataAssemblyForExcel
        # data = DataAssemblyForExcel(self.file_path, excel_requirement_metadata['signal_datas'], excel_requirement_metadata['msg_info'], excel_requirement_metadata['gw_mapping'], excel_requirement_metadata['route_info']).assemble_all()
        pass
        # return data


def get_can_routing_name(file_paths):
    """批量识别每个 Excel 的"网关路由表"sheet 名称。

    主要流程：
        - 对每个 Excel 文件打开工作簿，获取 sheetnames；
        - 通过 get_sheet_name_by_ai（结合 kb 的关键词/描述 + prompt）让模型选择路由表 sheet；
        - 返回列表，每项包含 file_path 与 routing_name。

    返回:
        List[{"routing_name": str, "file_path": str}]
    """
    res = []
    for file_path in file_paths:
        res1 = {
            "routing_name": "",
            "file_path": ""
        }
        wb = wbparser.parse_and_get_single_excel(file_path)
        if not wb:
            logging.error(f"Failed to parse Excel file: {file_path}")
            res.append(res1)
            continue
        sheet_names = wb.sheetnames
        input_prompt = prompts_data.get("QueryGwSheetNamePrompt",[])
        module = "can_routing"
        module_name = "网关路由表"
        gw_routing_name = get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names)
        res1["routing_name"] = gw_routing_name
        res1["file_path"] = file_path
        res.append(res1)
    return res


def get_nvm_sheet_name(file_paths):
    """批量识别每个 Excel 的"NVM/存储配置表"sheet 名称。

    主要流程：
        - 读取工作簿 sheetnames；
        - 用 get_sheet_name_by_ai（kb + prompt）选择最可能的 NVM 配置 sheet。

    返回:
        List[{"sheet_name": str, "file_path": str}]
    """
    res = []
    for file_path in file_paths:
        wb = wbparser.parse_and_get_single_excel(file_path)
        if not wb:
            logging.error(f"Failed to parse Excel file: {file_path}")
            res.append({"sheet_name": "", "file_path": file_path})
            continue
        sheet_names = wb.sheetnames
        input_prompt = prompts_data.get("QueryNvmSheetNamePrompt")
        module = "nvm_block"
        module_name = "存储配置表"
        sheet_name = get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names)
        res.append({"sheet_name": sheet_name, "file_path": file_path})
    return res


def get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names):
    """根据业务模块（module）与知识库信息，从 sheet_names 中挑选目标 sheet。

    这里的"AI"并不直接读表内容，而是：
        - 从 global_var.kb['sheet_info'][module] 取关键词（sheet_name_key_words）与描述（sheet_description）；
        - 注入到 input_prompt 中，形成更可控的提示词；
        - 调用 common.get_sheet_name，在候选 sheet_names 中做匹配/模型选择。

    参数:
        input_prompt: prompt 模板字符串（通常来自 prompts_data）。
        module: 业务模块 key（如 can_routing / nvm_block / lin_signal）。
        module_name: 业务模块中文名，用于提示或日志。
        sheet_names: 工作簿中所有 sheet 名称列表。

    返回:
        sheet_name: 识别到的 sheet 名称（或空）。
    """
    sheet_name_key_words = global_var.kb.get("sheet_info", {}).get(module, {}).get("sheet_name_key_words", [])
    sheet_name_key_words_str = '、'.join(sheet_name_key_words)
    sheet_description = global_var.kb.get("sheet_info", {}).get(module, {}).get("sheet_description", [])
    sheet_description_str = '。'.join(sheet_description) + "。"
    generate_prompt = input_prompt.format(
        sheet_name_key_words=sheet_name_key_words_str,
        sheet_description=sheet_description_str)
    sheet_name = get_sheet_name(sheet_names, generate_prompt, module_name)
    return sheet_name


def get_lin_signal_schedule_data(file_paths):
    """批量识别 LIN 信号表/调度表 sheet，并标注类型。

    主要流程：
        - 对每个 Excel，先识别 lin_signal（Lin 信号表）相关 sheet；
        - 再识别 lin_schedule（Lin 调度表）相关 sheet；
        - 返回统一列表，使用 sheet_type 区分：
            - 'lin'：信号表
            - 'schedule'：调度表

    返回:
        List[{"file_path": str, "sheet_names": Union[str, List[str]], "sheet_type": str}]
    """
    res = []
    for file_path in file_paths:
        workbook = wbparser.parse_and_get_single_excel(file_path)
        if not workbook:
            logging.error(f"Failed to parse Excel file: {file_path}")
            continue
        sheet_names = workbook.sheetnames
        input_prompt = prompts_data.get("LinSignalSheetNamePrompt")
        module = "lin_signal"
        module_name = "Lin信号表"
        lin_sheet_names = get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names)
        if lin_sheet_names:
            res.append({"file_path": file_path, "sheet_names": lin_sheet_names, 'sheet_type': 'lin'})
        input_prompt = prompts_data.get("LinScheduleSheetNamePrompt")
        module = "lin_schedule"
        module_name = "Lin调度表"
        schedule_sheet_names = get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names)
        if schedule_sheet_names:
            res.append({"file_path": file_path, "sheet_names": schedule_sheet_names, 'sheet_type': 'schedule'})
    return res


def get_lin_signal_schedule_data_old(file_paths):
    """旧版：通过字符串规则（非 AI）识别 LIN 信号表/调度表 sheet。

    规则要点：
        - 信号表：sheet 名包含 'LIN'，且不包含 'LIN Legend' / '调度' / 'schedule' 等排除词；
        - 调度表：sheet 名包含 '调度表' 或 'schedule'。

    返回结构与 get_lin_signal_schedule_data 相同。
    """
    res = []
    for file_path in file_paths:
        workbook = wbparser.parse_and_get_single_excel(file_path)
        sheet_names = workbook.sheetnames
        lin_sheet_names = []
        schedule_sheet_names = []
        for sheet_name in sheet_names:
            in_fixes = ['LIN']
            not_in_fixes = ['LIN Legend', "调度", 'schedule']
            for in_fix in in_fixes:
                if in_fix.lower() in sheet_name.lower() and not any(
                        fix.lower() in sheet_name.lower() for fix in not_in_fixes):
                    lin_sheet_names.append(sheet_name)
            schedule_fixes = ['调度表', 'schedule']
            for schedule_fix in schedule_fixes:
                if schedule_fix in sheet_name.lower():
                    schedule_sheet_names.append(sheet_name)
        if lin_sheet_names:
            res.append({"file_path": file_path, "sheet_names": lin_sheet_names, 'sheet_type': 'lin'})
        if schedule_sheet_names:
            res.append({"file_path": file_path, "sheet_names": schedule_sheet_names, 'sheet_type': 'schedule'})
    return res


def get_routing_msg(file_path, gw_routing_name):
    """对单个 Excel 的指定路由表 sheet 做结构识别，返回 routing_msg。

    参数:
        file_path: Excel 文件路径。
        gw_routing_name: 网关路由表 sheet 名称。

    返回:
        routing_msg: 由 parse_can_excel 生成的结构识别结果；若入参缺失返回空 dict。
    """
    if not (gw_routing_name and file_path):
        return {}
    routing_msg = parse_can_excel(file_path, gw_routing_name)
    return routing_msg


def get_can_signal_sheet_names(file_paths):
    """批量识别 CAN 信号表 sheet 名称（基于 kb + prompt）。

    主要流程：
        - 读取每个工作簿的 sheetnames；
        - 使用 get_sheet_name_by_ai 选择"CAN信号表"相关 sheet；
        - 返回 file_path + sheet_names。

    返回:
        List[{"file_path": str, "sheet_names": Union[str, List[str]]}]
    """
    res = []
    for file_path in file_paths:
        workbook = wbparser.parse_and_get_single_excel(file_path)
        if not workbook:
            logging.error(f"Failed to parse Excel file: {file_path}")
            continue
        sheet_names = workbook.sheetnames
        input_prompt = prompts_data.get("CanSignalSheetNamePrompt")
        module = "can_signal"
        module_name = "CAN信号表"
        check_sheet_names = get_sheet_name_by_ai(input_prompt, module, module_name, sheet_names)
        if check_sheet_names:
            res.append({"file_path": file_path, "sheet_names": check_sheet_names})
    return res


def get_can_signal_sheet_names_old(file_paths):
    """旧版：通过简单前缀规则识别 CAN 信号表 sheet。

    匹配规则（只要命中其一）：
        - 'Tx_' / 'Rx_' / 'DebugMessage' / '_CAN'

    返回:
        List[{"file_path": str, "sheet_names": List[str]}]
    """
    res = []
    for file_path in file_paths:
        workbook = wbparser.parse_and_get_single_excel(file_path)
        sheet_names = workbook.sheetnames
        check_sheet_names = []
        for sheet_name in sheet_names:
            prefixes = ['Tx_', 'Rx_', 'DebugMessage', '_CAN']
            for prefix in prefixes:
                if prefix.lower() in sheet_name.lower():
                    check_sheet_names.append(sheet_name)
                    break
        if check_sheet_names:
            res.append({"file_path": file_path, "sheet_names": check_sheet_names})
    return res


def get_schedule_data(file_path, sheet_names):
    """读取指定 sheet_names 的 LIN 调度表数据并调用调度解析。

    主要流程：
        - 逐个 sheet：handle_merged_cells 补齐合并单元格显示值；
        - clean_sheet_data 去掉全空行/列；
        - 组织为 [{sheet_name: sheet_data}, ...] 的 input_list；
        - 调用 schedule_parse_function 并发/批量解析，输出结构化调度信息。

    返回:
        schedule_parse_function 的结果（依实现而定）；若无有效输入返回 None。
    """
    parse_data = wbparser.parse_and_get_single_excel(file_path)
    # 初始化最终结果存储
    input_list = []
    for sheet_name in sheet_names:
        sheet = parse_data[sheet_name]
        cleaned_data = handle_merged_cells(sheet)
        sheet_data = clean_sheet_data(cleaned_data)
        input_list.append({sheet_name: sheet_data})
    if input_list:
        return schedule_parse_function(input_list)


def get_nvm_main_data(file_path, sheet_names):
    """读取 NVM 配置相关 sheet，并生成字段映射后的主数据。

    主要流程：
        - 对每个 sheet：处理合并单元格 + 清理空行/空列；
        - 截取左上角区域（get_header_input，max_row/max_column 更大以覆盖常见 NVM 表头）；
        - 调用 nvm_map_function 让模型推断"字段 -> 表头列"的映射与抽取；
        - nvm_mapping_fields_validate：若映射缺失比例过高（>80%）认为误识别，剔除该 sheet。

    返回:
        {"file_path": file_path, "main_data": <nvm_map_function 输出>}
    """
    parse_data = wbparser.parse_and_get_single_excel(file_path)
    # 初始化最终结果存储
    header_input = []
    for sheet_name in sheet_names:
        sheet = parse_data[sheet_name]
        cleaned_data = handle_merged_cells(sheet)
        sheet_data = clean_sheet_data(cleaned_data)
        header_input.append({sheet_name: get_header_input(sheet_data, max_column=50, max_row=30)})
    res = {'file_path': file_path, 'main_data': nvm_map_function(header_input)}
    nvm_mapping_fields_validate(res, sheet_names)
    return res


def nvm_mapping_fields_validate(res, sheet_names):
    """对 nvm_map_function 的映射结果做"误识别"过滤。

    判定逻辑：
        - 取每个 sheet 的 block.fields 映射项；
        - 统计 value 为空（没找到对应列/字段）的 key 数量；
        - 若空值比例超过 80%，认为这个 sheet 不是目标 NVM 表或表头识别失败，直接从 res['main_data'] 移除。

    目的:
        减少 AI/规则误判导致的错误结构化数据。
    """
    if res['main_data']:
        for sheet_name in sheet_names:
            if sheet_name in res['main_data']:
                mapping_data = res['main_data'][sheet_name].get("block", {}).get("fields", {})
                length = len(mapping_data)
                if length > 0:
                    not_find_keys = []
                    for key, value in mapping_data.items():
                        if not value:
                            not_find_keys.append(key)
                    # 没找到的比例如果超过80% 则认为当前sheet_name数据错误
                    if len(not_find_keys) / length > 0.8:
                        logging.error(
                            f"NVM common sheet {sheet_name} not found fields: {not_find_keys} not found, so pop it")
                        res['main_data'].pop(sheet_name)


def get_signal_mapping_data(file_path, sheet_names, excel_signal_prompt):
    """对指定 sheet_names 做"信号字段 -> 表头列名"的映射生成，并返回列号信息。

    主要流程：
        1) 针对每个 sheet，抽取左上角表头区域（get_header_input），用于让模型识别表头结构；
        2) table_header_map_function 识别每个 sheet 的表头方向与行号（direction/index）；
        3) 对于横向表头（horizontal）：取最后一行作为表头，生成 column_mapping（表头名->列字母）与 header_index（表头行索引）；
        4) 将表头字符串交给 map_function + excel_signal_prompt 生成字段映射；
        5) signal_mapping_fields_validate：若映射缺失比例过高（>80%）则剔除该 sheet。

    返回:
        {
          "mapping_info": <map_function 输出>,
          "column_mapping": {sheet_name: {header: col_letter}},
          "header_index": {sheet_name: header_row_index_0_based}
        }

    备注:
        该函数包含大量 print 调试输出（可能用于离线验证），不影响核心返回结构。
    """
    parse_data = wbparser.parse_and_get_single_excel(file_path)

    # 初始化最终结果存储
    input_list = []
    header_input = []
    for sheet_name in sheet_names:
        sheet = parse_data[sheet_name]
        # print(f"sheet:{sheet}")
        cleaned_data = handle_merged_cells(sheet)
        # print(f"cleaned_data:{cleaned_data}")
        sheet_data = clean_sheet_data(cleaned_data)
        # print(f"sheet_data:{sheet_data}")
        header_input.append({sheet_name: get_header_input(sheet_data)})


    # print("="*50)
    # print(f"heand:{header_input}")
    # print()
    header_infos = table_header_map_function(header_input)
    # print("="*50)
    # print(header_infos)
    column_mapping = {}
    header_index = {}
    for sheet_name in sheet_names:
        sheet = parse_data[sheet_name]
        cleaned_data = handle_merged_cells(sheet)
        sheet_data = clean_sheet_data(cleaned_data)

        # 提取当前Sheet的表头
        header_info = header_infos.get(sheet_name)
        if header_info and header_info['direction'] == 'horizontal':
            # 取最后一行
            index = header_info['index'][-1]  # 取最后一行
            # 表头
            header = sheet_data[index - 1]
            # 表头对应的列号
            column_mapping[sheet_name] = get_excel_column_mapping(header)
            # 表头所在的行号
            header_index[sheet_name] = index - 1
            input_list.append({sheet_name: ",".join(header)})  # 当前Sheet的表头

    if input_list:
        mapping = map_function(input_list, excel_signal_prompt)
        signal_mapping_fields_validate(mapping, sheet_names)
        return {"mapping_info": mapping, "column_mapping": column_mapping, "header_index": header_index}
    return {}


def signal_mapping_fields_validate(mapping, sheet_names):
    """对信号字段映射结果做"误识别"过滤。

    判定逻辑：
        - 检查 mapping[sheet_name]["ComSignal"] 下每个字段的映射；
        - 若缺少 title_name 或 title_name 为空，记为"未找到"；
        - 未找到比例 > 80% 则认为该 sheet 识别错误，从 mapping 中剔除。

    参数:
        mapping: map_function 输出的映射结果（会被原地修改）。
        sheet_names: 参与处理的 sheet 名列表。
    """
    for sheet_name in sheet_names:
        if sheet_name in mapping:
            mapping_data = mapping.get(sheet_name)["ComSignal"]
            length = len(mapping_data)
            if length > 0:
                not_find_keys = []
                for key, value in mapping_data.items():
                    if "title_name" not in value or not value["title_name"]:
                        not_find_keys.append(key)
                # 没找到的比例如果超过80% 则认为当前sheet_name数据错误
                if len(not_find_keys) / length > 0.8:
                    logging.error(f"signal sheet {sheet_name} not found fields: {not_find_keys} not found, so pop it")
                    mapping.pop(sheet_name)

if __name__ == '__main__':
    # file_path = r"D:\KoTEI_CODE\Agent_Autosar_Backend\new_data\AY5-G Project_CMX_EEA3.0_CCU_LIN_MIX_V1.4 - 增加诊断(LIN诊断相关).xlsx"
    # sheet_names = ["LIN1", "LIN2"]
    #
    # sys_prompt = prompts_data.get("LinExcelSignalPrompt")
    # ans = get_signal_mapping_data(file_path, sheet_names, sys_prompt)
    # print(ans)
    # quit()

    header_input = [{'LIN1': [['CCU LIN1 Messages', '', '', '', '', '', '', '', '', ''], ['ECU (Tx)', 'Frame Name', 'LIN ID (hex)', 'Protected ID (hex)', 'Frame Length (bytes)', 'Frame Send Type', 'Frame Cycle Time (ms)', 'Signal Name', 'Signal Comment', 'Start Bit Position'], ['Application Frames', '', '', '', '', '', '', '', '', ''], ['CCURT1', 'CCU_LIN1_1', '0x10', '0x50', 8, 'UF', 60, 'UMM_UsageModeSt', 'Usage mode state', 0], ['CCURT2', 'CCU_LIN1_1', '0x10', '0x50', 8, 'UF', 60, 'WW_ScreenType', 'Type of windscreen', 3], ['CCURT1', 'CCU_LIN1_1', '0x10', '0x50', 8, 'UF', 60, 'Not used', '', 7], ['CCU', 'CCU_LIN1_1', '0x10', '0x50', 8, 'UF', 60, 'BCS_VehSpdVD', 'Quality/fault information to vehicle speed', 12], ['CCU', 'CCU_LIN1_1', '0x10', '0x50', 8, 'UF', 60, 'BCS_VehSpd', 'Vehicle speed', 13], ['CCURT2', 'CCU_LIN1_1', '0x10', '0x50', 8, 'UF', 60, 'WW_RS_Cali_Data', 'Rain sensor calibration data', 26], ['CCURT1', 'CCU_LIN1_1', '0x10', '0x50', 8, 'UF', 60, 'Not used', '', 30]]}, {'LIN2': [['CCU LIN2 Messages', '', '', '', '', '', '', '', '', ''], ['ECU (Tx)', 'Frame Name', 'LIN ID (hex)', 'Protected ID (hex)', 'Frame Length (bytes)', 'Frame Send Type', 'Frame Cycle Time (ms)', 'Signal Name', 'Signal Comment', 'Start Bit Position'], ['Application Frames', '', '', '', '', '', '', '', '', ''], ['CCURT1', 'CCU_LIN2_1', '0x11', '0x11', 8, 'UF', 60, 'UMM_UsageModeSt', 'Usage mode state', 0], ['CCURT1', 'CCU_LIN2_1', '0x11', '0x11', 8, 'UF', 60, 'UMM_PowerKeepMode', 'Power Keep Mode', 3], ['CCURT1', 'CCU_LIN2_1', '0x11', '0x11', 8, 'UF', 60, 'UMM_PowerOffEnable', 'Power off Enable Status', 4], ['CCURT1', 'CCU_LIN2_1', '0x11', '0x11', 8, 'UF', 60, 'UMM_PowerKeepEnable', 'Power keep Enable Status', 5], ['CCURT1', 'CCU_LIN2_1', '0x11', '0x11', 8, 'UF', 60, 'Not used', '', 6], ['CCURT1', 'CCU_LIN2_1', '0x11', '0x11', 8, 'UF', 60, 'VMM_VehModeSt', 'Vehicle mode state', 8], ['CCURT1', 'CCU_LIN2_1', '0x11', '0x11', 8, 'UF', 60, 'VMM_LastVehModeSt', 'Last time vehicle mode', 12]]}]
    header_infos = table_header_map_function(header_input)
    print(header_infos)